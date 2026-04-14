"""Tests for geoview_cpt.parsers.field_book — Phase A-2 A2.2c."""
from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import pytest

from geoview_cpt.parsers.field_book import (
    FieldBookEntry,
    FieldBookParseError,
    FieldBookTable,
    detect_field_book,
    parse_field_book,
)


# ---------------------------------------------------------------------------
# Synthetic 야장 factory
# ---------------------------------------------------------------------------


def _make_field_book(tmp_path: Path, *, rows: int = 3) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header rows 2-3 — leave as None padding but populate a few cells
    ws.cell(2, 2, "Day")
    ws.cell(2, 3, "Point")
    ws.cell(2, 5, "수심")
    ws.cell(2, 6, "관입심도")
    ws.cell(3, 6, "CPT(m)")
    ws.cell(3, 7, "PC(m)")
    ws.cell(3, 11, "X")
    ws.cell(3, 12, "Y")

    templates = [
        {
            "day": datetime(2025, 9, 1),
            "cpt": "CPT1",
            "pc": "PC2",
            "wd": "88m",
            "cpt_m": 2.03,
            "pc_m": None,
            "dp": time(10, 0),
            "seat": time(10, 30),
            "end": time(12, 30),
            "x": 523112.67,
            "y": 4520000.0,
        },
        {
            "day": datetime(2025, 9, 2),
            "cpt": "CPT4-1",
            "pc": "PC5-1",
            "wd": "112m",
            "cpt_m": 1.06,
            "pc_m": None,
            "dp": time(16, 51),
            "seat": time(17, 10),
            "end": time(18, 0),
            "x": 537925.4,
            "y": 4525000.0,
        },
        {
            "day": datetime(2025, 9, 3),
            "cpt": "CPT4-1",
            "pc": "PC5-1",
            "wd": None,
            "cpt_m": 2,
            "pc_m": "-",
            "dp": time(12, 20),
            "seat": time(12, 40),
            "end": time(13, 20),
            "x": None,
            "y": None,
        },
    ]

    for i, tpl in enumerate(templates[:rows]):
        r = 4 + i
        ws.cell(r, 2, tpl["day"])
        ws.cell(r, 3, tpl["cpt"])
        ws.cell(r, 4, tpl["pc"])
        ws.cell(r, 5, tpl["wd"])
        ws.cell(r, 6, tpl["cpt_m"])
        ws.cell(r, 7, tpl["pc_m"])
        ws.cell(r, 8, tpl["dp"])
        ws.cell(r, 9, tpl["seat"])
        ws.cell(r, 10, tpl["end"])
        ws.cell(r, 11, tpl["x"])
        ws.cell(r, 12, tpl["y"])

    out = tmp_path / "synth_fieldbook.xlsx"
    wb.save(out)
    wb.close()
    return out


@pytest.fixture
def synth_field_book(tmp_path) -> Path:
    return _make_field_book(tmp_path)


# ---------------------------------------------------------------------------
# Entry dataclass
# ---------------------------------------------------------------------------


class TestFieldBookEntry:
    def test_is_empty_detection(self):
        empty = FieldBookEntry(row=5)
        assert empty.is_empty is True
        populated = FieldBookEntry(row=5, cpt_point="CPT1")
        assert populated.is_empty is False

    def test_day_coerced_from_datetime(self):
        e = FieldBookEntry(row=1, day=datetime(2025, 9, 1))
        assert e.day == date(2025, 9, 1)


# ---------------------------------------------------------------------------
# Parse — synthetic
# ---------------------------------------------------------------------------


class TestParseSynthetic:
    def test_returns_table(self, synth_field_book):
        t = parse_field_book(synth_field_book)
        assert isinstance(t, FieldBookTable)
        assert t.sheet_name == "Sheet1"

    def test_three_entries(self, synth_field_book):
        t = parse_field_book(synth_field_book)
        assert len(t) == 3

    def test_first_entry_fields(self, synth_field_book):
        t = parse_field_book(synth_field_book)
        e = t.entries[0]
        assert e.row == 4
        assert e.day == date(2025, 9, 1)
        assert e.cpt_point == "CPT1"
        assert e.pc_point == "PC2"
        assert e.water_depth_text == "88m"
        assert e.water_depth_m == 88.0
        assert e.cpt_depth_m == pytest.approx(2.03)
        assert e.dp_start_time == time(10, 0)
        assert e.seabed_contact_time == time(10, 30)
        assert e.end_time == time(12, 30)
        assert e.cpt_x == pytest.approx(523112.67)

    def test_none_water_depth(self, synth_field_book):
        t = parse_field_book(synth_field_book)
        third = t.entries[2]
        assert third.water_depth_text == ""
        assert third.water_depth_m is None

    def test_distinct_points_listing(self, synth_field_book):
        t = parse_field_book(synth_field_book)
        # CPT4-1 appears twice but should be listed once
        assert t.cpt_points() == ["CPT1", "CPT4-1"]
        assert t.pc_points() == ["PC2", "PC5-1"]

    def test_header_raw_captured(self, synth_field_book):
        t = parse_field_book(synth_field_book)
        assert len(t.header_raw) == 3
        # One of the header cells should contain our Korean marker
        flat = [c for row in t.header_raw for c in row if isinstance(c, str)]
        assert any("수심" in s for s in flat)

    def test_raw_row_preserved(self, synth_field_book):
        t = parse_field_book(synth_field_book)
        e = t.entries[0]
        assert len(e.raw_row) >= 12
        # col 3 (1-indexed) should hold the CPT point
        assert e.raw_row[2] == "CPT1"


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------


class TestErrors:
    def test_missing_file(self, tmp_path):
        with pytest.raises(FieldBookParseError, match="not found"):
            parse_field_book(tmp_path / "nope.xlsx")

    def test_too_few_rows(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "x")
        out = tmp_path / "tiny.xlsx"
        wb.save(out)
        wb.close()
        with pytest.raises(FieldBookParseError, match="not enough rows"):
            parse_field_book(out)


# ---------------------------------------------------------------------------
# Detect
# ---------------------------------------------------------------------------


class TestDetect:
    def test_positive(self, synth_field_book):
        assert detect_field_book(synth_field_book) is True

    def test_missing_file(self, tmp_path):
        assert detect_field_book(tmp_path / "nope.xlsx") is False

    def test_wrong_extension(self, tmp_path):
        p = tmp_path / "x.txt"
        p.write_text("hi")
        assert detect_field_book(p) is False


# ---------------------------------------------------------------------------
# Real JAKO 야장 (optional)
# ---------------------------------------------------------------------------


_REAL_FB = Path(r"H:/자코/JAKO_Korea_area/Excel_변환_data/JAKO_Korea_CPT 야장.xlsx")
fb_required = pytest.mark.skipif(
    not _REAL_FB.exists(),
    reason="JAKO 야장 real sample not mounted (H: drive)",
)


@pytest.fixture(scope="session")
def real_field_book():
    if not _REAL_FB.exists():
        pytest.skip("JAKO 야장 real sample not mounted")
    return parse_field_book(_REAL_FB)


@fb_required
class TestRealFieldBook:
    def test_reads_real_jako_log(self, real_field_book):
        t = real_field_book
        assert len(t) > 5
        points = t.cpt_points()
        assert any("CPT2" in p for p in points)
        assert any("CPT9" in p for p in points)

    def test_real_water_depth_parsed(self, real_field_book):
        parsed = [e.water_depth_m for e in real_field_book if e.water_depth_m is not None]
        assert len(parsed) > 0

    def test_real_coords_present(self, real_field_book):
        x_vals = [e.cpt_x for e in real_field_book if e.cpt_x is not None]
        assert len(x_vals) > 0
        for x in x_vals:
            assert 500_000 < x < 600_000
