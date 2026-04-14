"""Tests for geoview_cpt.parsers.excel_yw — Phase A-2 A2.2a."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import numpy as np
import pytest

from geoview_cpt.model import CPTSounding
from geoview_cpt.parsers.excel_yw import (
    YwParseError,
    YwParseOptions,
    detect_yw_xlsx,
    parse_yw_xlsx,
)


# ---------------------------------------------------------------------------
# Synthetic YW xlsx factory
# ---------------------------------------------------------------------------


def _make_yw_xlsx(
    tmp_path: Path,
    sheet_name: str = "DATA YW-99",
    *,
    rows: int = 20,
    missing_header: bool = False,
    empty_data: bool = False,
) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if missing_header:
        ws.cell(1, 1, "Something Else")
        ws.cell(1, 2, "Not a cone")
    else:
        ws.cell(1, 1, "Test length[1]")
        ws.cell(1, 2, "Measured cone resistance[2]")
        ws.cell(1, 3, "Local friction[3]")
        ws.cell(1, 4, "Pore pressure u2[6]")
        ws.cell(2, 1, "m")
        ws.cell(2, 2, "MPa")
        ws.cell(2, 3, "MPa")
        ws.cell(2, 4, "MPa")
    if not empty_data:
        for i in range(rows):
            depth = round(0.02 * (i + 1), 3)
            qc = round(0.001 * (i + 1), 4)
            fs = round(0.0005 * (i + 1), 4)
            u2 = round(0.0002 * (i + 1), 4)
            ws.cell(3 + i, 1, depth)
            ws.cell(3 + i, 2, qc)
            ws.cell(3 + i, 3, fs)
            ws.cell(3 + i, 4, u2)
    out = tmp_path / "synth_yw.xlsx"
    wb.save(out)
    wb.close()
    return out


@pytest.fixture
def synth_yw_path(tmp_path) -> Path:
    return _make_yw_xlsx(tmp_path)


# ---------------------------------------------------------------------------
# Happy path
# ---------------------------------------------------------------------------


class TestParseYwSynthetic:
    def test_returns_cpt_sounding(self, synth_yw_path):
        s = parse_yw_xlsx(synth_yw_path)
        assert isinstance(s, CPTSounding)

    def test_sounding_id_from_sheet(self, synth_yw_path):
        s = parse_yw_xlsx(synth_yw_path)
        assert s.name == "YW-99"
        assert s.header is not None
        assert s.header.sounding_id == "YW-99"

    def test_channels_populated(self, synth_yw_path):
        s = parse_yw_xlsx(synth_yw_path)
        assert set(s.channels.keys()) == {"depth", "qc", "fs", "u2"}
        assert len(s.channels["depth"]) == 20

    def test_unit_conversion_kpa(self, synth_yw_path):
        s = parse_yw_xlsx(synth_yw_path)
        # fs is 0.0005 MPa at first row → 0.5 kPa after conversion
        assert s.channels["fs"].unit == "kPa"
        assert s.channels["u2"].unit == "kPa"
        assert s.channels["fs"].values[0] == pytest.approx(0.5)

    def test_unit_conversion_disabled(self, tmp_path):
        path = _make_yw_xlsx(tmp_path)
        s = parse_yw_xlsx(
            path,
            options=YwParseOptions(convert_friction_to_kpa=False, convert_pore_to_kpa=False),
        )
        assert s.channels["fs"].unit == "MPa"
        assert s.channels["u2"].unit == "MPa"

    def test_helms_equipment_defaults(self, synth_yw_path):
        s = parse_yw_xlsx(synth_yw_path)
        assert s.header.cone_base_area_mm2 == 1000.0
        assert s.header.cone_area_ratio_a == pytest.approx(0.71)
        assert s.header.partner_name == "HELMS"
        assert s.header.sounding_type == "PCPT"
        assert s.header.max_push_depth_m is not None

    def test_max_depth_stored(self, synth_yw_path):
        s = parse_yw_xlsx(synth_yw_path)
        assert s.max_depth_m == pytest.approx(0.4)

    def test_metadata_captured(self, synth_yw_path):
        s = parse_yw_xlsx(synth_yw_path)
        meta = s.metadata
        assert meta["source_format"] == "yw_xlsx"
        assert meta["sheet_name"] == "DATA YW-99"
        assert "column_map" in meta


# ---------------------------------------------------------------------------
# Error paths
# ---------------------------------------------------------------------------


class TestErrors:
    def test_missing_file(self, tmp_path):
        with pytest.raises(YwParseError, match="not found"):
            parse_yw_xlsx(tmp_path / "nope.xlsx")

    def test_missing_headers_raises(self, tmp_path):
        path = _make_yw_xlsx(tmp_path, missing_header=True)
        with pytest.raises(YwParseError, match="missing expected headers"):
            parse_yw_xlsx(path)

    def test_empty_data_raises(self, tmp_path):
        path = _make_yw_xlsx(tmp_path, empty_data=True)
        with pytest.raises(YwParseError, match="no data rows"):
            parse_yw_xlsx(path)


# ---------------------------------------------------------------------------
# Detect
# ---------------------------------------------------------------------------


class TestDetect:
    def test_positive(self, synth_yw_path):
        assert detect_yw_xlsx(synth_yw_path) is True

    def test_missing_file(self, tmp_path):
        assert detect_yw_xlsx(tmp_path / "nope.xlsx") is False

    def test_wrong_extension(self, tmp_path):
        p = tmp_path / "x.txt"
        p.write_text("hi")
        assert detect_yw_xlsx(p) is False

    def test_wrong_header(self, tmp_path):
        path = _make_yw_xlsx(tmp_path, missing_header=True)
        assert detect_yw_xlsx(path) is False


# ---------------------------------------------------------------------------
# Real YW file (optional)
# ---------------------------------------------------------------------------


_REAL_YW = Path(
    r"H:/야월해상풍력단지 지반조사 용역 결과보고서_rev7/CPT 데이터 분석/Raw_Data/YW-01 (PCPT).xlsx"
)
yw_required = pytest.mark.skipif(
    not _REAL_YW.exists(),
    reason="YW real sample not mounted (H: drive)",
)


@pytest.fixture(scope="session")
def real_yw_sounding():
    if not _REAL_YW.exists():
        pytest.skip("YW real sample not mounted")
    return parse_yw_xlsx(_REAL_YW)


@yw_required
class TestRealYw:
    def test_reads_yw_01(self, real_yw_sounding):
        s = real_yw_sounding
        assert s.header.sounding_id.startswith("YW")
        assert len(s.channels["depth"]) > 1000
        assert s.max_depth_m > 10

    def test_channel_values_reasonable(self, real_yw_sounding):
        s = real_yw_sounding
        qc = s.channels["qc"].values
        # Cone resistance typical range 0..80 MPa
        assert qc.min() >= -1
        assert qc.max() < 200
        depth = s.channels["depth"].values
        # Monotonically non-decreasing for a field survey
        assert np.all(np.diff(depth) >= -1e-6)
