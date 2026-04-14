"""
geoview_cpt.parsers.field_book
================================
JAKO acquisition 야장 (field book) reader (Phase A-2 A2.2c).

Wave 0 reconnaissance found the 야장 to be a single-sheet ``.xlsx``
that logs one CPT/PC pair per row. Column 1 is always empty; data
columns run from column 2 through column 18 in fixed positions:

    col  2  Day                 (date)
    col  3  CPT Point           (str, e.g. "CPT4-1")
    col  4  PC Point            (str, e.g. "PC5-1")
    col  5  Water depth         (str with "m", e.g. "88m")
    col  6  CPT depth (m)
    col  7  PC depth (m)
    col  8  DP start time       (time)
    col  9  Seabed contact time (time) — 안착
    col 10  End time            (time) — 종료
    col 11  CPT X               (float)
    col 12  CPT Y               (float)
    col 13  PC X                (float)
    col 14  PC Y                (float)
    col 15  Attempt count CPT   (int/str)
    col 16  Attempt count PC    (int/str)
    col 17  Grade               (str) — 등급
    col 18  Analysis depth      (str/float) — 분석심도

Real JAKO Korea 야장 has rows 4-25 populated. Some soundings span
multiple rows (retries for rigging / tidal window issues) — the reader
captures each row as one :class:`FieldBookEntry` and does not attempt
to fold retries under a single parent.

The header text is CP949-encoded in the source file. openpyxl returns
it as mojibake; rather than guess the encoding we rely on **column
position** for recognition. The header rows are captured raw into
:attr:`FieldBookTable.header_raw` so downstream tools (B3.8) can show
them in the original encoding if the user has the right codepage.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date as date_cls, datetime, time
from pathlib import Path
from typing import Any

__all__ = [
    "FieldBookEntry",
    "FieldBookTable",
    "FieldBookParseError",
    "parse_field_book",
    "detect_field_book",
]


class FieldBookParseError(Exception):
    """Raised when a 야장 ``.xlsx`` file cannot be parsed."""


@dataclass
class FieldBookEntry:
    """One row of the JAKO 야장 log."""

    row: int                                # source-row index (1-based)
    day: date_cls | None = None
    cpt_point: str = ""
    pc_point: str = ""
    water_depth_text: str = ""              # "88m" — kept verbatim
    water_depth_m: float | None = None      # parsed float when possible
    cpt_depth_m: float | None = None
    pc_depth_m: float | None = None
    dp_start_time: time | None = None
    seabed_contact_time: time | None = None
    end_time: time | None = None
    cpt_x: float | None = None
    cpt_y: float | None = None
    pc_x: float | None = None
    pc_y: float | None = None
    cpt_attempt_count: int | None = None
    pc_attempt_count: int | None = None
    grade: str = ""
    analysis_depth: str = ""
    raw_row: list[Any] = field(default_factory=list)   # preserve untouched cells

    def __post_init__(self) -> None:
        if isinstance(self.day, datetime):
            self.day = self.day.date()

    @property
    def is_empty(self) -> bool:
        """True when the row has no identifying point and no timestamps."""
        return not (
            self.cpt_point
            or self.pc_point
            or self.dp_start_time
            or self.end_time
            or self.day
        )


@dataclass
class FieldBookTable:
    """Collection of :class:`FieldBookEntry` plus raw header rows."""

    source_path: Path
    sheet_name: str
    header_raw: list[list[Any]] = field(default_factory=list)
    entries: list[FieldBookEntry] = field(default_factory=list)

    def __len__(self) -> int:
        return len(self.entries)

    def __iter__(self):
        return iter(self.entries)

    def cpt_points(self) -> list[str]:
        """Return the distinct CPT point labels in source order."""
        seen: list[str] = []
        for e in self.entries:
            if e.cpt_point and e.cpt_point not in seen and e.cpt_point != "-":
                seen.append(e.cpt_point)
        return seen

    def pc_points(self) -> list[str]:
        seen: list[str] = []
        for e in self.entries:
            if e.pc_point and e.pc_point not in seen and e.pc_point != "-":
                seen.append(e.pc_point)
        return seen


# ---------------------------------------------------------------------------
# Detect
# ---------------------------------------------------------------------------


_MIN_ROWS = 5
_MIN_COLS = 10


def detect_field_book(path: Path | str) -> bool:
    """Return True when ``path`` looks like a JAKO 야장 ``.xlsx``."""
    p = Path(path)
    if p.suffix.lower() != ".xlsx":
        return False
    try:
        import openpyxl
    except ImportError:
        return False
    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    except Exception:
        return False
    try:
        if not wb.sheetnames:
            return False
        ws = wb[wb.sheetnames[0]]
        if ws.max_row < _MIN_ROWS or ws.max_column < _MIN_COLS:
            return False
        # A 야장 always has a date in column 2 of one of rows 4..10
        for r in range(4, min(ws.max_row + 1, 11)):
            day = ws.cell(r, 2).value
            if isinstance(day, (datetime, date_cls)):
                return True
        return False
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------


def parse_field_book(path: Path | str) -> FieldBookTable:
    """Read a JAKO 야장 ``.xlsx`` into a :class:`FieldBookTable`."""
    p = Path(path)
    if not p.exists():
        raise FieldBookParseError(f"file not found: {p}")

    try:
        import openpyxl
    except ImportError as exc:
        raise FieldBookParseError(f"openpyxl not installed: {exc}") from exc

    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    except Exception as exc:
        raise FieldBookParseError(f"cannot open {p}: {exc}") from exc

    try:
        if not wb.sheetnames:
            raise FieldBookParseError(f"{p} has no sheets")
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        if ws.max_row < _MIN_ROWS:
            raise FieldBookParseError(f"{p}: not enough rows to contain a 야장")

        header_raw = [
            [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 35))]
            for r in range(1, 4)
        ]

        entries: list[FieldBookEntry] = []
        for r in range(4, ws.max_row + 1):
            raw_row = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 35))]
            entry = _row_to_entry(r, raw_row)
            if entry.is_empty:
                continue
            entries.append(entry)
    finally:
        wb.close()

    return FieldBookTable(
        source_path=p,
        sheet_name=sheet_name,
        header_raw=header_raw,
        entries=entries,
    )


def _row_to_entry(row_idx: int, raw: list[Any]) -> FieldBookEntry:
    """Map a raw cell list (col 1 at index 0) to a :class:`FieldBookEntry`."""

    def cell(col_1b: int):
        """Return cell at 1-based column, or None when out of range."""
        idx = col_1b - 1
        return raw[idx] if 0 <= idx < len(raw) else None

    wd_text, wd_value = _split_water_depth(cell(5))
    return FieldBookEntry(
        row=row_idx,
        day=_coerce_date(cell(2)),
        cpt_point=_coerce_str(cell(3)),
        pc_point=_coerce_str(cell(4)),
        water_depth_text=wd_text,
        water_depth_m=wd_value,
        cpt_depth_m=_coerce_float_or_none(cell(6)),
        pc_depth_m=_coerce_float_or_none(cell(7)),
        dp_start_time=_coerce_time(cell(8)),
        seabed_contact_time=_coerce_time(cell(9)),
        end_time=_coerce_time(cell(10)),
        cpt_x=_coerce_float_or_none(cell(11)),
        cpt_y=_coerce_float_or_none(cell(12)),
        pc_x=_coerce_float_or_none(cell(13)),
        pc_y=_coerce_float_or_none(cell(14)),
        cpt_attempt_count=_coerce_int_or_none(cell(15)),
        pc_attempt_count=_coerce_int_or_none(cell(16)),
        grade=_coerce_str(cell(17)),
        analysis_depth=_coerce_str(cell(18)),
        raw_row=list(raw),
    )


# ---------------------------------------------------------------------------
# coercion helpers
# ---------------------------------------------------------------------------


def _coerce_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_date(value: Any) -> date_cls | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_cls):
        return value
    return None


def _coerce_time(value: Any) -> time | None:
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    return None


def _coerce_float_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        txt = value.strip().rstrip("m").strip()
        if not txt or txt == "-":
            return None
        try:
            return float(txt)
        except ValueError:
            return None
    return None


def _coerce_int_or_none(value: Any) -> int | None:
    f = _coerce_float_or_none(value)
    return int(f) if f is not None else None


def _split_water_depth(value: Any) -> tuple[str, float | None]:
    """Parse ``'88m'`` → ``('88m', 88.0)``. Missing cells → ``('', None)``."""
    if value is None:
        return "", None
    text = str(value).strip()
    if not text:
        return "", None
    number = _coerce_float_or_none(text)
    return text, number
