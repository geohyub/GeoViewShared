"""
geoview_cpt.parsers.excel_jako
================================
JAKO Korea vendor ``.xls`` reader (Phase A-2 A2.2b).

Wave 0 reconnaissance found the JAKO format — produced by the Gouda
WISON onboard acquisition software, version tag ``6.98:1`` — to be a
single-sheet ``.xls`` with a fixed metadata header block followed by a
12-column data table:

    r0          "Software Version"                  (label)
    r1          "6.98:1"                            (version)
    r3          ┌ Project Name │ Client Name │ Location │ Vessel │ Client
                │ Operator    │ Cone        │ Notes
    r4          └ (values)
    r6          ┌ Fix Number │ Water Depth │ Push Name │ Test Number
                │ Raw Data File │ Computed Data File │ Max Tip │ Max Incline
    r7          └ (values)
    r9          ┌ Tip Area (mm) │ N Value
    r10         └ (values, e.g. 200 │ 15)
    r12         "Tip Area Factor = 0.7032"
    r13         "Hydrostatic Pressure (MPa) = 0.87"

    r15         Date&Time │ Pen (m) │ Tip (MPa Qc) │ Cu (MPa) │ Sleeve (MPa)
                │ Measured Pore (MPa) │ TiltX │ TiltY │ Combined Tilt
                │ Altimeter (m) │ Voltage (V) │ Current (A)
    r16..       data samples

Equipment defaults read directly from the header block:
``cone_base_area_mm2`` from ``Tip Area (mm)`` and ``cone_area_ratio_a``
from the ``Tip Area Factor`` text line. Both are overridable via
:class:`JakoParseOptions`.

Canonical channel mapping:

    depth  ← Pen (m)                        (m)
    qc     ← Tip (MPa Qc)                   (MPa)
    fs     ← Sleeve (MPa) × 1000            (kPa)
    u2     ← Measured Pore (MPa) × 1000     (kPa)
    incl   ← Combined Tilt (Degrees)        (deg)
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np

from geoview_cpt.model import CPTChannel, CPTHeader, CPTSounding

__all__ = [
    "JakoParseOptions",
    "JakoParseError",
    "parse_jako_xls",
    "detect_jako_xls",
]


class JakoParseError(Exception):
    """Raised when a JAKO ``.xls`` file cannot be parsed."""


@dataclass
class JakoParseOptions:
    partner_name: str = "Geoview"
    cone_base_area_mm2_default: float = 200.0
    cone_area_ratio_a_default: float = 0.7032
    equipment_vendor: str = "Gouda Geo-Equipment"
    equipment_model: str = "WISON-APB"


# ---------------------------------------------------------------------------


_DATA_HEADER_MARKERS = ("date&time", "pen", "tip")
_TIP_FACTOR_RE = re.compile(r"tip\s*area\s*factor\s*=\s*([0-9.]+)", re.IGNORECASE)
_HYDRO_RE = re.compile(r"hydrostatic\s*pressure.*?([0-9.+-]+)", re.IGNORECASE)
_JAKO_TIMESTAMP_RE = re.compile(r"#(\d{2})/(\d{2})/(\d{4})\s+(\d{2}):(\d{2}):(\d{2})#")


def detect_jako_xls(path: Path | str) -> bool:
    """Return True when ``path`` looks like a JAKO vendor export (``.xls``/``.xlsx``)."""
    p = Path(path)
    if p.suffix.lower() not in (".xls", ".xlsx"):
        return False
    try:
        sh = _open_sheet(p)
    except Exception:
        return False
    if sh.nrows < 16:
        return False
    probe = " ".join(str(sh.cell_value(r, 0)) for r in range(min(sh.nrows, 5))).lower()
    return "software version" in probe


def parse_jako_xls(
    path: Path | str,
    *,
    options: JakoParseOptions | None = None,
) -> CPTSounding:
    """
    Read a JAKO vendor Excel export into a :class:`CPTSounding`.

    Supports both the native ``.xls`` produced by the Gouda WISON
    acquisition software and ``.xlsx`` copies (handy for tests — the
    cell layout is identical).
    """
    opts = options or JakoParseOptions()
    p = Path(path)
    if not p.exists():
        raise JakoParseError(f"file not found: {p}")

    try:
        sh = _open_sheet(p)
    except ModuleNotFoundError as exc:
        raise JakoParseError(f"required reader missing: {exc}") from exc
    except Exception as exc:
        raise JakoParseError(f"cannot open {p}: {exc}") from exc

    meta = _read_meta_block(sh)
    data_header_row = _find_data_header(sh)
    channels = _read_channels(sh, data_header_row)

    if channels["depth"].size == 0:
        raise JakoParseError(f"{p} contains no data rows")

    sounding_id = (
        meta.get("push_name")
        or meta.get("project_name")
        or p.stem
    )

    cone_area = _float_or(meta.get("tip_area_mm2"), opts.cone_base_area_mm2_default)
    area_ratio = _float_or(meta.get("tip_area_factor"), opts.cone_area_ratio_a_default)

    header = CPTHeader(
        sounding_id=sounding_id,
        project_name=str(meta.get("project_name", "")),
        client=str(meta.get("client_name", meta.get("client", ""))),
        partner_name=opts.partner_name,
        operator=str(meta.get("operator", "")),
        vessel=str(meta.get("vessel", "")),
        water_depth_m=_float_or_none(meta.get("water_depth")),
        sounding_type="PCPT",
        equipment_vendor=opts.equipment_vendor,
        equipment_model=opts.equipment_model,
        cone_base_area_mm2=cone_area,
        cone_area_ratio_a=area_ratio,
        max_push_depth_m=float(channels["depth"].max()),
        started_at=channels.get("first_timestamp"),
        completed_at=channels.get("last_timestamp"),
    )

    sounding = CPTSounding(
        handle=0,
        element_tag="",
        name=sounding_id,
        file_name=p.name,
        input_count=int(channels["depth"].size),
        output_count=int(channels["depth"].size),
        max_depth_m=float(channels["depth"].max()),
        unit_system=0,
    )
    sounding.header = header
    sounding.channels = {
        "depth": CPTChannel(name="depth", unit="m", values=channels["depth"]),
        "qc":    CPTChannel(name="qc",    unit="MPa", values=channels["qc"]),
        "fs":    CPTChannel(name="fs",    unit="kPa", values=channels["fs"] * 1000.0),
        "u2":    CPTChannel(name="u2",    unit="kPa", values=channels["u2"] * 1000.0),
        "incl":  CPTChannel(name="incl",  unit="deg", values=channels["incl"]),
    }
    sounding.metadata.update(
        {
            "source_format": "jako_xls",
            "source_path": str(p),
            "software_version": meta.get("software_version", ""),
            "raw_meta": meta,
            "data_header_row": data_header_row,
        }
    )
    return sounding


# ---------------------------------------------------------------------------
# sheet adapter — isolates xlrd vs openpyxl
# ---------------------------------------------------------------------------


class _SheetView:
    """Minimal sheet interface shared by both backends."""

    def __init__(self, nrows: int, ncols: int, cell_fn):
        self.nrows = nrows
        self.ncols = ncols
        self._cell = cell_fn

    def cell_value(self, r: int, c: int):
        return self._cell(r, c)


def _open_sheet(path: Path) -> _SheetView:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        import xlrd

        wb = xlrd.open_workbook(str(path))
        sh = wb.sheet_by_index(0)
        return _SheetView(
            nrows=sh.nrows,
            ncols=sh.ncols,
            cell_fn=sh.cell_value,
        )
    if suffix == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        # Materialize into a list so we can close the workbook safely.
        nrows = ws.max_row
        ncols = ws.max_column
        grid = [
            [ws.cell(r, c).value for c in range(1, ncols + 1)]
            for r in range(1, nrows + 1)
        ]
        wb.close()

        def cell_fn(r: int, c: int):
            if 0 <= r < nrows and 0 <= c < ncols:
                v = grid[r][c]
                return v if v is not None else ""
            return ""

        return _SheetView(nrows=nrows, ncols=ncols, cell_fn=cell_fn)
    raise JakoParseError(f"unsupported extension {suffix!r} for JAKO reader")


# ---------------------------------------------------------------------------
# meta block
# ---------------------------------------------------------------------------


def _read_meta_block(sh) -> dict[str, Any]:
    """Scan the first ~15 rows for label/value pairs."""
    meta: dict[str, Any] = {}

    # Software version
    for r in range(min(sh.nrows, 3)):
        label = str(sh.cell_value(r, 0)).strip()
        if label.lower() == "software version" and r + 1 < sh.nrows:
            meta["software_version"] = str(sh.cell_value(r + 1, 0)).strip()
            break

    # Label rows r3/r6/r9 + value rows r4/r7/r10 (approx)
    for header_r, value_r in ((3, 4), (6, 7), (9, 10)):
        if value_r >= sh.nrows:
            continue
        for c in range(min(sh.ncols, 12)):
            label = str(sh.cell_value(header_r, c)).strip()
            if not label:
                continue
            value = sh.cell_value(value_r, c)
            key = _normalize_key(label)
            if key:
                meta[key] = value

    # r12/r13 free-text labels
    for r in range(11, min(sh.nrows, 16)):
        cell = str(sh.cell_value(r, 0))
        if not cell:
            continue
        m = _TIP_FACTOR_RE.search(cell)
        if m:
            meta["tip_area_factor"] = float(m.group(1))
        m = _HYDRO_RE.search(cell)
        if m:
            try:
                meta["hydrostatic_pressure_mpa"] = float(m.group(1))
            except ValueError:
                pass

    return meta


def _normalize_key(label: str) -> str:
    s = label.strip().lower()
    s = re.sub(r"[()]", "", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    # Common remaps for downstream clarity
    remap = {
        "tip_area_mm": "tip_area_mm2",
        "water_depth": "water_depth",
        "max_tip_mpa": "max_tip_mpa",
        "max_incline_degrees": "max_incline_deg",
    }
    return remap.get(s, s)


# ---------------------------------------------------------------------------
# data table
# ---------------------------------------------------------------------------


def _find_data_header(sh) -> int:
    for r in range(min(sh.nrows, 30)):
        row_txt = " ".join(str(sh.cell_value(r, c)).lower() for c in range(min(sh.ncols, 4)))
        if all(marker in row_txt for marker in _DATA_HEADER_MARKERS):
            return r
    raise JakoParseError("data header row (Date&Time / Pen / Tip) not found")


def _read_channels(sh, header_row: int) -> dict[str, Any]:
    """Extract canonical channels starting at ``header_row + 1``."""
    header_cells = [
        str(sh.cell_value(header_row, c)) for c in range(sh.ncols)
    ]
    col_of = {}
    for idx, cell in enumerate(header_cells):
        low = cell.lower()
        if "date" in low and "time" in low and "time" not in col_of:
            col_of["time"] = idx
        elif "pen" in low and "depth" not in col_of:
            col_of["depth"] = idx
        elif "tip" in low and "qc" in low and "qc" not in col_of:
            col_of["qc"] = idx
        elif "sleeve" in low and "fs" not in col_of:
            col_of["fs"] = idx
        elif "pore" in low and "u2" not in col_of:
            col_of["u2"] = idx
        elif "combined" in low and "incl" not in col_of:
            col_of["incl"] = idx

    required = {"depth", "qc", "fs", "u2"}
    missing = required - set(col_of)
    if missing:
        raise JakoParseError(
            f"JAKO data header missing required columns: {missing} — got {header_cells!r}"
        )

    depths: list[float] = []
    qcs: list[float] = []
    fss: list[float] = []
    u2s: list[float] = []
    incls: list[float] = []
    first_ts: datetime | None = None
    last_ts: datetime | None = None

    for r in range(header_row + 1, sh.nrows):
        d_cell = sh.cell_value(r, col_of["depth"])
        try:
            d_val = float(d_cell)
        except (TypeError, ValueError):
            continue
        depths.append(d_val)
        qcs.append(_coerce_float(sh.cell_value(r, col_of["qc"])))
        fss.append(_coerce_float(sh.cell_value(r, col_of["fs"])))
        u2s.append(_coerce_float(sh.cell_value(r, col_of["u2"])))
        incls.append(
            _coerce_float(sh.cell_value(r, col_of["incl"])) if "incl" in col_of else 0.0
        )
        if "time" in col_of:
            ts = _parse_timestamp(sh.cell_value(r, col_of["time"]))
            if ts is not None:
                if first_ts is None:
                    first_ts = ts
                last_ts = ts

    return {
        "depth": np.asarray(depths, dtype=np.float64),
        "qc": np.asarray(qcs, dtype=np.float64),
        "fs": np.asarray(fss, dtype=np.float64),
        "u2": np.asarray(u2s, dtype=np.float64),
        "incl": np.asarray(incls, dtype=np.float64),
        "first_timestamp": first_ts,
        "last_timestamp": last_ts,
    }


def _coerce_float(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _float_or(value, default: float) -> float:
    try:
        return float(value) if value not in (None, "") else default
    except (TypeError, ValueError):
        return default


def _float_or_none(value) -> float | None:
    try:
        return float(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _parse_timestamp(value) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    m = _JAKO_TIMESTAMP_RE.match(s)
    if not m:
        return None
    dd, mm, yyyy, hh, mi, ss = (int(g) for g in m.groups())
    try:
        return datetime(yyyy, mm, dd, hh, mi, ss)
    except ValueError:
        return None
