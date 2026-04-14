"""
geoview_cpt.parsers.excel_yw
================================
HELMS Yawol CPT Excel template reader (Phase A-2 A2.2a).

Wave 0 reconnaissance found the YW template to be a single-sheet
``.xlsx`` with a fixed four-column data block:

    row 1       header                              row 2       unit
    ─────────   ────────────────────────────────    ─────────   ────
    col 1       Test length[1]                      m
    col 2       Measured cone resistance[2]         MPa
    col 3       Local friction[3]                   MPa
    col 4       Pore pressure u2[6]                 MPa

Data samples start at row 3. Real YW-01 carries 1,672 rows of data
(max test length ≈ 33.4 m). Sheet name follows ``DATA YW-01`` — this
reader extracts the sounding identifier from that name.

Equipment defaults: **HELMS 1000 mm² cone, area ratio 0.71**
(master plan §5.1, Wave 0 helms_doosan_profile). The caller can
override via :attr:`YwParseOptions`.

Consumer contract: returns a fully-formed :class:`CPTSounding` with
:class:`CPTHeader` populated, raw channels ``depth``, ``qc``, ``fs``
(converted to kPa), ``u2`` (converted to kPa). Derived channels are
left empty — those land in A2.5.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import numpy as np

from geoview_cpt.model import CPTChannel, CPTHeader, CPTSounding

__all__ = [
    "YwParseOptions",
    "YwParseError",
    "parse_yw_xlsx",
    "detect_yw_xlsx",
]


class YwParseError(Exception):
    """Raised when a YW ``.xlsx`` file cannot be parsed."""


@dataclass
class YwParseOptions:
    """Runtime knobs for the YW reader."""

    partner_name: str = "HELMS"
    vessel: str = ""
    cone_base_area_mm2: float = 1000.0
    cone_area_ratio_a: float = 0.71
    equipment_vendor: str = "Geomarine"
    equipment_model: str = ""
    convert_friction_to_kpa: bool = True
    convert_pore_to_kpa: bool = True


# ---------------------------------------------------------------------------
# Header recognition
# ---------------------------------------------------------------------------


_HEADER_FIELDS = {
    "depth": re.compile(r"test\s*length", re.IGNORECASE),
    "qc":    re.compile(r"measured\s*cone\s*resistance", re.IGNORECASE),
    "fs":    re.compile(r"local\s*friction", re.IGNORECASE),
    "u2":    re.compile(r"pore\s*pressure", re.IGNORECASE),
}

_SHEET_RE = re.compile(r"data\s*(?P<id>YW[-\s]*\w+)", re.IGNORECASE)


def _normalize_sounding_id(sheet_name: str) -> str:
    m = _SHEET_RE.match(sheet_name.strip())
    if m:
        raw = m.group("id").upper()
        return raw.replace(" ", "-")
    return sheet_name.strip().upper()


def _match_columns(header_row: list) -> dict[str, int]:
    """Map canonical names (``depth``, ``qc``, ``fs``, ``u2``) to 0-based col index."""
    out: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for name, pat in _HEADER_FIELDS.items():
            if name in out:
                continue
            if pat.search(cell):
                out[name] = idx
    missing = [n for n in _HEADER_FIELDS if n not in out]
    if missing:
        raise YwParseError(
            f"YW sheet is missing expected headers: {missing} — got row={header_row!r}"
        )
    return out


# ---------------------------------------------------------------------------
# Detect
# ---------------------------------------------------------------------------


def detect_yw_xlsx(path: Path | str) -> bool:
    """Return True when ``path`` looks like a YW CPT Excel export."""
    p = Path(path)
    if p.suffix.lower() != ".xlsx":
        return False
    try:
        import openpyxl
    except ImportError as exc:
        raise YwParseError(f"openpyxl not installed: {exc}") from exc
    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    except Exception:
        return False
    try:
        if not wb.sheetnames:
            return False
        first = wb.sheetnames[0]
        if not first.lower().startswith("data"):
            return False
        ws = wb[first]
        header_row = [ws.cell(1, c).value for c in range(1, min(ws.max_column + 1, 10))]
        try:
            _match_columns(header_row)
        except YwParseError:
            return False
        return True
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------


def parse_yw_xlsx(
    path: Path | str,
    *,
    options: YwParseOptions | None = None,
) -> CPTSounding:
    """
    Read a YW template ``.xlsx`` and return a populated :class:`CPTSounding`.

    Args:
        path:    filesystem path to a ``YW-XX (PCPT).xlsx`` file.
        options: runtime knobs (partner name, cone geometry, unit conversion).

    Raises:
        YwParseError: on missing/unreadable files, openpyxl failures,
                      header mismatches, or empty data blocks.
    """
    opts = options or YwParseOptions()
    p = Path(path)
    if not p.exists():
        raise YwParseError(f"file not found: {p}")

    try:
        import openpyxl
    except ImportError as exc:
        raise YwParseError(f"openpyxl not installed: {exc}") from exc

    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    except Exception as exc:
        raise YwParseError(f"cannot open workbook {p}: {exc}") from exc

    try:
        if not wb.sheetnames:
            raise YwParseError(f"{p} has no sheets")
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = list(next(row_iter))
        except StopIteration:
            raise YwParseError(f"{p} is empty")
        col_map = _match_columns(header_row)
        try:
            unit_row = list(next(row_iter))
        except StopIteration:
            raise YwParseError(f"{p} has header row only")

        depth, qc, fs, u2 = _read_data_stream(row_iter, col_map)
    finally:
        wb.close()

    if depth.size == 0:
        raise YwParseError(f"{p} contains no data rows")

    sounding_id = _normalize_sounding_id(sheet_name)

    # Unit conversion: YW ships fs and u2 in MPa — convert to kPa to match
    # the canonical channel convention used by geoview_cpt.derivation.
    fs_kpa = fs * 1000.0 if opts.convert_friction_to_kpa else fs
    u2_kpa = u2 * 1000.0 if opts.convert_pore_to_kpa else u2

    header = CPTHeader(
        sounding_id=sounding_id,
        partner_name=opts.partner_name,
        vessel=opts.vessel,
        cone_base_area_mm2=opts.cone_base_area_mm2,
        cone_area_ratio_a=opts.cone_area_ratio_a,
        equipment_vendor=opts.equipment_vendor,
        equipment_model=opts.equipment_model,
        sounding_type="PCPT",
    )
    header.max_push_depth_m = float(depth.max())

    sounding = CPTSounding(
        handle=0,
        element_tag="",
        name=sounding_id,
        file_name=p.name,
        input_count=int(depth.size),
        output_count=int(depth.size),
        max_depth_m=float(depth.max()),
        unit_system=0,
    )
    sounding.header = header
    sounding.channels = {
        "depth": CPTChannel(name="depth", unit="m", values=depth),
        "qc":    CPTChannel(name="qc",    unit="MPa", values=qc),
        "fs":    CPTChannel(name="fs",    unit="kPa" if opts.convert_friction_to_kpa else "MPa", values=fs_kpa),
        "u2":    CPTChannel(name="u2",    unit="kPa" if opts.convert_pore_to_kpa else "MPa", values=u2_kpa),
    }
    sounding.metadata.update(
        {
            "source_format": "yw_xlsx",
            "source_path": str(p),
            "sheet_name": sheet_name,
            "column_map": col_map,
            "header_row": header_row,
            "unit_row": unit_row,
        }
    )
    return sounding


def _read_data_stream(rows_iter, col_map: dict[str, int]) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Drain ``rows_iter`` (already past the header rows) into numpy arrays."""
    depth_idx = col_map["depth"]
    qc_idx = col_map["qc"]
    fs_idx = col_map["fs"]
    u2_idx = col_map["u2"]

    depths: list[float] = []
    qcs: list[float] = []
    fss: list[float] = []
    u2s: list[float] = []

    for row in rows_iter:
        if row is None:
            continue
        if depth_idx >= len(row):
            continue
        d = row[depth_idx]
        if d is None:
            continue
        try:
            d_val = float(d)
        except (TypeError, ValueError):
            continue
        depths.append(d_val)
        qcs.append(_coerce_float(row[qc_idx] if qc_idx < len(row) else None))
        fss.append(_coerce_float(row[fs_idx] if fs_idx < len(row) else None))
        u2s.append(_coerce_float(row[u2_idx] if u2_idx < len(row) else None))

    return (
        np.asarray(depths, dtype=np.float64),
        np.asarray(qcs, dtype=np.float64),
        np.asarray(fss, dtype=np.float64),
        np.asarray(u2s, dtype=np.float64),
    )


def _coerce_float(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0
