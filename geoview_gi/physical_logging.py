"""
geoview_gi.physical_logging
================================
Density + PS-wave (P-S suspension) logging data model + xlsx readers.

Wave 0 3rd-round recon confirmed the 지오플러스이엔지 (GeoPlus) 2025-05
deliverable ships as:

 - ``(지오플러스이엔지)...밀도검층DATASHEET.xlsx``
     3 sheets (YW-1-0, YW-4, YW-12), per sheet:
       col A  DEPTH (m)
       col B  LSD (CPS)    — long-spaced density count
       col C  DENSITY (g/cm³)
       col E  depth range (text, "0.01~9.0")
       col F  stratum label (Korean)
       col G  mean density per stratum

 - ``(지오플러스이엔지)...음파검층DATASHEET.xlsx``
     3 sheets (YW-1-0, YW-4, YW-12), per sheet:
       col A  Depth (EL.m)
       col B  Depth (GL.-m)
       col C  Rock Type (Korean)
       col D  Vp (km/s)
       col E  Vs (km/s)
       col F  γ (kN/m³)
       col G  Gd — dynamic shear modulus (MPa)
       col H  Ed — dynamic Young's modulus (MPa)
       col I  Kd — dynamic bulk modulus (MPa)
       col J  ud — dynamic Poisson's ratio

Everything above col G on the PS-wave sheet is *precomputed* by the
vendor, so the :class:`PSWaveLog` dataclass stores them verbatim and
our dynamic-modulus helpers serve as **cross-check** functions in the
test suite (not the source of truth).

Q29 status: HWP originals exist but are not parsed. v1 consumes the
xlsx DATASHEET exports directly; v1.1 can optionally ingest HWP via
``pyhwp``.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import numpy as np

__all__ = [
    "DensityStratum",
    "DensityLog",
    "PSWaveLog",
    "PhysicalLoggingParseError",
    "parse_density_datasheet_xlsx",
    "parse_ps_wave_datasheet_xlsx",
    "dynamic_shear_modulus",
    "dynamic_young_modulus",
    "dynamic_poisson_ratio",
]


class PhysicalLoggingParseError(Exception):
    """Raised when a density / PS-wave datasheet cannot be read."""


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------


@dataclass
class DensityStratum:
    """One stratum from the "mean density per layer" sidebar table."""

    top_m: float
    base_m: float
    stratum_label: str = ""
    mean_density_g_cm3: float | None = None


@dataclass
class DensityLog:
    """
    Long-spaced density (LSD) log — one borehole, one continuous profile.

    Attributes:
        borehole_id:    short name extracted from the sheet name
                        (e.g. "YW-1-0", "YW-4", "YW-12").
        sheet_name:     original sheet name (preserved for round-trip / UI).
        depth_m:        depth array (m).
        lsd_cps:        raw LSD count (counts per second).
        density_g_cm3:  bulk density (g/cm³).
        strata:         optional list of :class:`DensityStratum` parsed
                        from the vendor's mean-density side table.
        method:         acquisition method label (default HELMS default).
        source_path:    originating xlsx on disk.
    """

    borehole_id: str
    sheet_name: str
    depth_m: np.ndarray
    lsd_cps: np.ndarray
    density_g_cm3: np.ndarray
    strata: list[DensityStratum] = field(default_factory=list)
    method: str = "Long-spaced Density Log"
    source_path: Path | None = None

    def __len__(self) -> int:
        return int(self.depth_m.shape[0])

    @property
    def mean_density_g_cm3(self) -> float:
        if self.density_g_cm3.size == 0:
            return float("nan")
        return float(np.nanmean(self.density_g_cm3))


@dataclass
class PSWaveLog:
    """
    P-S suspension (seismic) log — velocities + vendor-precomputed moduli.

    Cross-check properties (``shear_modulus_cross_check`` etc.) recompute
    from (Vp, Vs, γ) — tests assert these match vendor columns to within
    rounding, proving the vendor output is internally consistent.
    """

    borehole_id: str
    sheet_name: str
    depth_el_m: np.ndarray            # Depth (EL.m)
    depth_gl_m: np.ndarray            # Depth (GL.-m)
    rock_type: list[str]
    vp_km_s: np.ndarray
    vs_km_s: np.ndarray
    gamma_kn_m3: np.ndarray
    gd_vendor_mpa: np.ndarray         # Gd column (vendor precomputed)
    ed_vendor_mpa: np.ndarray         # Ed column
    kd_vendor_mpa: np.ndarray         # Kd column
    poisson_vendor: np.ndarray        # ud column
    casing_depth_m: float = 0.0
    source_path: Path | None = None

    def __len__(self) -> int:
        return int(self.depth_gl_m.shape[0])

    @property
    def shear_modulus_cross_check_mpa(self) -> np.ndarray:
        return dynamic_shear_modulus(self.vp_km_s, self.vs_km_s, self.gamma_kn_m3)

    @property
    def young_modulus_cross_check_mpa(self) -> np.ndarray:
        return dynamic_young_modulus(self.vp_km_s, self.vs_km_s, self.gamma_kn_m3)

    @property
    def poisson_cross_check(self) -> np.ndarray:
        return dynamic_poisson_ratio(self.vp_km_s, self.vs_km_s)


# ---------------------------------------------------------------------------
# Dynamic moduli — cross-check helpers (not the source of truth)
# ---------------------------------------------------------------------------


# g = 9.81 m/s² used to convert γ (kN/m³) → mass density (kg/m³)
_G_MS2 = 9.81


def _mass_density_kg_m3(gamma_kn_m3: np.ndarray) -> np.ndarray:
    """Convert unit weight γ (kN/m³) to mass density ρ (kg/m³)."""
    return np.asarray(gamma_kn_m3, dtype=np.float64) * 1000.0 / _G_MS2


def dynamic_shear_modulus(
    vp_km_s: np.ndarray,
    vs_km_s: np.ndarray,
    gamma_kn_m3: np.ndarray,
) -> np.ndarray:
    """
    G_dyn = ρ × Vs² — returned in MPa.

    Vs input is km/s so we multiply by 1000 to get m/s before squaring.
    ``G_dyn = ρ (kg/m³) × Vs² (m²/s²)`` → Pa. Divide by 10⁶ for MPa.
    """
    rho = _mass_density_kg_m3(gamma_kn_m3)
    vs_ms = np.asarray(vs_km_s, dtype=np.float64) * 1000.0
    g_pa = rho * vs_ms * vs_ms
    return g_pa / 1e6


def dynamic_young_modulus(
    vp_km_s: np.ndarray,
    vs_km_s: np.ndarray,
    gamma_kn_m3: np.ndarray,
) -> np.ndarray:
    """
    E_dyn = 2·G·(1 + ν_dyn) where ν_dyn is computed from (Vp, Vs).
    Result in MPa.
    """
    g = dynamic_shear_modulus(vp_km_s, vs_km_s, gamma_kn_m3)
    nu = dynamic_poisson_ratio(vp_km_s, vs_km_s)
    return 2.0 * g * (1.0 + nu)


def dynamic_poisson_ratio(
    vp_km_s: np.ndarray, vs_km_s: np.ndarray
) -> np.ndarray:
    """
    ν_dyn = (Vp² − 2·Vs²) / (2·(Vp² − Vs²))
    Dimensionally consistent regardless of velocity unit (both Vp, Vs in
    km/s here). Clamped to < 0.5 by the denominator guard.
    """
    vp2 = np.asarray(vp_km_s, dtype=np.float64) ** 2
    vs2 = np.asarray(vs_km_s, dtype=np.float64) ** 2
    denom = 2.0 * (vp2 - vs2)
    denom = np.where(np.abs(denom) > 1e-9, denom, 1e-9)
    return (vp2 - 2.0 * vs2) / denom


# ---------------------------------------------------------------------------
# xlsx readers — GeoPlus DATASHEET format
# ---------------------------------------------------------------------------


_SHEET_ID_RE = re.compile(r"(YW[-\s]*\d+(?:-\d+)?)", re.IGNORECASE)


def _borehole_id_from_sheet(sheet_name: str) -> str:
    m = _SHEET_ID_RE.search(sheet_name or "")
    if m:
        return m.group(1).upper().replace(" ", "-")
    # Fall back to the raw sheet name (mojibake-safe — caller can ignore)
    return sheet_name


def _as_float_array(values: list[Any]) -> np.ndarray:
    out = np.empty(len(values), dtype=np.float64)
    for i, v in enumerate(values):
        try:
            out[i] = float(v) if v is not None else np.nan
        except (TypeError, ValueError):
            out[i] = np.nan
    return out


def parse_density_datasheet_xlsx(path: Path | str) -> list[DensityLog]:
    """
    Read a density DATASHEET xlsx into a list of :class:`DensityLog`.

    One :class:`DensityLog` per sheet; sheets are expected to match the
    per-borehole convention (one sheet → one borehole).
    """
    import openpyxl

    p = Path(path)
    if not p.exists():
        raise PhysicalLoggingParseError(f"file not found: {p}")
    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    except Exception as exc:
        raise PhysicalLoggingParseError(f"cannot open {p}: {exc}") from exc

    logs: list[DensityLog] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            log = _parse_density_sheet(ws, sheet_name)
            log.source_path = p
            logs.append(log)
    finally:
        wb.close()
    if not logs:
        raise PhysicalLoggingParseError(f"{p}: no sheets produced a DensityLog")
    return logs


def _parse_density_sheet(ws, sheet_name: str) -> DensityLog:
    # Primary table: cols 1-3 (DEPTH, LSD, DENSITY), data from row 2
    depths: list[Any] = []
    lsd: list[Any] = []
    dens: list[Any] = []
    rows = ws.iter_rows(min_row=2, values_only=True)
    # Side table: cols 5-7 (DEPTH range text, stratum label, mean density)
    side_entries: list[DensityStratum] = []
    for row in rows:
        if len(row) >= 3 and row[0] is not None:
            d = row[0]
            if isinstance(d, (int, float)):
                depths.append(d)
                lsd.append(row[1])
                dens.append(row[2])
        if len(row) >= 7 and row[4] is not None:
            depth_range = row[4]
            label = row[5] or ""
            mean_dens = row[6]
            top_m, base_m = _parse_depth_range(depth_range)
            if top_m is not None and base_m is not None:
                side_entries.append(
                    DensityStratum(
                        top_m=top_m,
                        base_m=base_m,
                        stratum_label=str(label).strip(),
                        mean_density_g_cm3=(
                            float(mean_dens) if isinstance(mean_dens, (int, float)) else None
                        ),
                    )
                )
    return DensityLog(
        borehole_id=_borehole_id_from_sheet(sheet_name),
        sheet_name=sheet_name,
        depth_m=_as_float_array(depths),
        lsd_cps=_as_float_array(lsd),
        density_g_cm3=_as_float_array(dens),
        strata=side_entries,
    )


def _parse_depth_range(text: Any) -> tuple[float | None, float | None]:
    if not isinstance(text, str):
        return None, None
    m = re.match(r"\s*([\d.]+)\s*~\s*([\d.]+)\s*", text)
    if not m:
        return None, None
    try:
        return float(m.group(1)), float(m.group(2))
    except ValueError:
        return None, None


def parse_ps_wave_datasheet_xlsx(path: Path | str) -> list[PSWaveLog]:
    """Read a PS-wave DATASHEET xlsx into a list of :class:`PSWaveLog`."""
    import openpyxl

    p = Path(path)
    if not p.exists():
        raise PhysicalLoggingParseError(f"file not found: {p}")
    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    except Exception as exc:
        raise PhysicalLoggingParseError(f"cannot open {p}: {exc}") from exc

    logs: list[PSWaveLog] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            log = _parse_ps_wave_sheet(ws, sheet_name)
            log.source_path = p
            logs.append(log)
    finally:
        wb.close()
    if not logs:
        raise PhysicalLoggingParseError(f"{p}: no sheets produced a PSWaveLog")
    return logs


def _parse_ps_wave_sheet(ws, sheet_name: str) -> PSWaveLog:
    # Header at row 2 (0-indexed), data from row 4 onwards. Blank row at 3.
    depth_el: list[Any] = []
    depth_gl: list[Any] = []
    rock: list[str] = []
    vp: list[Any] = []
    vs: list[Any] = []
    gamma: list[Any] = []
    gd: list[Any] = []
    ed: list[Any] = []
    kd: list[Any] = []
    nu: list[Any] = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None or not isinstance(row[0], (int, float)):
            continue
        depth_el.append(row[0])
        depth_gl.append(row[1] if len(row) > 1 else None)
        rock.append(str(row[2]) if len(row) > 2 and row[2] is not None else "")
        vp.append(row[3] if len(row) > 3 else None)
        vs.append(row[4] if len(row) > 4 else None)
        gamma.append(row[5] if len(row) > 5 else None)
        gd.append(row[6] if len(row) > 6 else None)
        ed.append(row[7] if len(row) > 7 else None)
        kd.append(row[8] if len(row) > 8 else None)
        nu.append(row[9] if len(row) > 9 else None)

    return PSWaveLog(
        borehole_id=_borehole_id_from_sheet(sheet_name),
        sheet_name=sheet_name,
        depth_el_m=_as_float_array(depth_el),
        depth_gl_m=_as_float_array(depth_gl),
        rock_type=rock,
        vp_km_s=_as_float_array(vp),
        vs_km_s=_as_float_array(vs),
        gamma_kn_m3=_as_float_array(gamma),
        gd_vendor_mpa=_as_float_array(gd),
        ed_vendor_mpa=_as_float_array(ed),
        kd_vendor_mpa=_as_float_array(kd),
        poisson_vendor=_as_float_array(nu),
    )
