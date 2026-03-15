"""
GeoView Excel Writer
====================
Professional Excel report generation with GeoView branding.

Usage:
    from geoview_common.reporting.excel_writer import ExcelReport

    report = ExcelReport("Survey Calculation Report")
    report.add_header_row(["Parameter", "Value", "Unit", "Status"])
    report.add_data_row(["Water Depth", 45.2, "m", "OK"])
    report.save("output.xlsx")
"""

from pathlib import Path
from datetime import datetime
from typing import List, Optional, Union

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..styles import colors


class ExcelReport:
    """Professional Excel report with GeoView branding."""

    def __init__(self, title: str, author: str = "GeoView",
                 project: str = "", sheet_name: str = "Report"):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required: pip install openpyxl")

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name
        self.title = title
        self.author = author
        self.project = project
        self._row = 1

        # Define reusable styles
        self._header_font = Font(name="Pretendard", size=9, bold=True,
                                 color="FFFFFF")
        self._header_fill = PatternFill(start_color="1E3A5F",
                                        end_color="1E3A5F",
                                        fill_type="solid")
        self._body_font = Font(name="Pretendard", size=9)
        self._result_font = Font(name="Pretendard", size=9, bold=True,
                                  color="2E7D32")
        self._warn_font = Font(name="Pretendard", size=9, bold=True,
                                color="E53E3E")
        self._even_fill = PatternFill(start_color="F0F4F8",
                                       end_color="F0F4F8",
                                       fill_type="solid")
        self._good_fill = PatternFill(start_color="E8F5E9",
                                       end_color="E8F5E9",
                                       fill_type="solid")
        self._warn_fill = PatternFill(start_color="FFE0B2",
                                       end_color="FFE0B2",
                                       fill_type="solid")
        self._danger_fill = PatternFill(start_color="FFCDD2",
                                         end_color="FFCDD2",
                                         fill_type="solid")
        self._thin_border = Border(
            left=Side(style="thin", color="CBD5E0"),
            right=Side(style="thin", color="CBD5E0"),
            top=Side(style="thin", color="CBD5E0"),
            bottom=Side(style="thin", color="CBD5E0"),
        )
        self._center = Alignment(horizontal="center", vertical="center")

        self._write_title_block()

    def _write_title_block(self):
        """Write title, project info, and date."""
        # Title row (merged)
        title_font = Font(name="Pretendard", size=14, bold=True, color="1E3A5F")
        self.ws.merge_cells(start_row=1, start_column=1,
                            end_row=1, end_column=6)
        cell = self.ws.cell(row=1, column=1, value=self.title)
        cell.font = title_font

        # Info row
        info_font = Font(name="Pretendard", size=9, color="4A5568")
        info_parts = []
        if self.project:
            info_parts.append(f"Project: {self.project}")
        info_parts.append(f"Author: {self.author}")
        info_parts.append(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        info_text = "  |  ".join(info_parts)

        self.ws.merge_cells(start_row=2, start_column=1,
                            end_row=2, end_column=6)
        cell = self.ws.cell(row=2, column=1, value=info_text)
        cell.font = info_font

        self._row = 4  # Leave a blank row

    def add_section(self, title: str):
        """Add a section header row."""
        section_font = Font(name="Pretendard", size=10, bold=True,
                            color="1E3A5F")
        section_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7",
                                    fill_type="solid")
        self.ws.merge_cells(start_row=self._row, start_column=1,
                            end_row=self._row, end_column=6)
        cell = self.ws.cell(row=self._row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        self._row += 1

    def add_header_row(self, headers: List[str],
                       widths: Optional[List[int]] = None):
        """Add a column header row with navy background."""
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self._row, column=col_idx, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._center
            cell.border = self._thin_border
            if widths and col_idx <= len(widths):
                self.ws.column_dimensions[
                    get_column_letter(col_idx)].width = widths[col_idx - 1]
        self._row += 1

    def add_data_row(self, values: list, status: Optional[str] = None):
        """Add a data row. status: 'good', 'warn', 'danger', or None."""
        is_even = (self._row % 2 == 0)
        fill = None
        if status == "good":
            fill = self._good_fill
        elif status == "warn":
            fill = self._warn_fill
        elif status == "danger":
            fill = self._danger_fill
        elif is_even:
            fill = self._even_fill

        for col_idx, value in enumerate(values, 1):
            cell = self.ws.cell(row=self._row, column=col_idx, value=value)
            cell.font = self._body_font
            cell.border = self._thin_border
            cell.alignment = self._center
            if fill:
                cell.fill = fill

            # Format numbers
            if isinstance(value, float):
                cell.number_format = "0.000" if abs(value) < 1 else "0.00"
        self._row += 1

    def add_blank_row(self):
        """Add an empty row for spacing."""
        self._row += 1

    def auto_fit_columns(self, min_width: int = 10, max_width: int = 40):
        """Auto-fit column widths based on content."""
        for col in self.ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            width = min(max(max_len + 2, min_width), max_width)
            self.ws.column_dimensions[col_letter].width = width

    def save(self, filepath: Union[str, Path]):
        """Save the workbook."""
        self.auto_fit_columns()
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(filepath))
        return filepath
