"""
GeoView QC — Unified Report Builder
=====================================
Generates cross-domain QC reports combining MAG, Sonar, and Seismic results.

Supports:
- Excel (openpyxl) — Multi-sheet workbook with per-domain tabs
- Word (python-docx via design_system) — Professional report with GeoView branding
- PDF (reportlab) — Standalone PDF with charts

Copyright (c) 2025-2026 Geoview Co., Ltd.
"""

from __future__ import annotations

import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from .models import QCDomain, QCResult, QCProjectSummary, QCStatus, QCGrade

# Ensure shared lib
_shared = Path(__file__).resolve().parents[3]
if str(_shared) not in sys.path:
    sys.path.insert(0, str(_shared))


# ---------------------------------------------------------------------------
# Shared constants
# ---------------------------------------------------------------------------

_NAVY = "1E3A5F"
_WHITE = "FFFFFF"
_BORDER_COLOR = "CBD5E0"
_ALT_ROW_1 = "F7FAFC"
_ALT_ROW_2 = "FFFFFF"
_FONT = "Pretendard"
_FONT_FALLBACK = "Calibri"

# Grade fill colors (openpyxl hex, no #)
_GRADE_FILLS = {
    QCGrade.A: ("C6F6D5", "276749"),  # light green bg, dark green text
    QCGrade.B: ("BEE3F8", "2A4365"),  # light blue bg, dark blue text
    QCGrade.C: ("FEEBC8", "7B341E"),  # light orange bg, dark orange text
    QCGrade.D: ("FED7D7", "9B2C2C"),  # light red bg, dark red text
    QCGrade.F: ("E2E8F0", "4A5568"),  # light gray bg, dark gray text
}

# Score font colors
_SCORE_COLORS = {
    "pass": "38A169",   # green
    "warn": "ED8936",   # orange
    "fail": "E53E3E",   # red
}


def _score_tier(score: float) -> str:
    if score >= 80:
        return "pass"
    elif score >= 50:
        return "warn"
    return "fail"


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

def generate_excel_report(
    summary: QCProjectSummary,
    output_path: Optional[str | Path] = None,
) -> BytesIO | Path:
    """Generate unified Excel QC report.

    Creates a workbook with:
    - Summary sheet (project info banner, overall KPIs)
    - Per-domain sheets (MAG, Sonar, Seismic) with result tables
    - Professional styling: navy headers, alternating rows, grade fills,
      auto-width columns, thin borders
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Reusable styles ----
    navy_fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
    accent_fill = PatternFill(start_color="2D5F8A", end_color="2D5F8A", fill_type="solid")
    white_font = Font(name=_FONT, size=10, color=_WHITE, bold=True)
    header_font = Font(name=_FONT, size=10, bold=True, color=_NAVY)
    body_font = Font(name=_FONT, size=10)
    muted_font = Font(name=_FONT, size=9, color="718096")
    title_font = Font(name=_FONT, size=16, bold=True, color=_WHITE)
    subtitle_font = Font(name=_FONT, size=11, color="A0AEC0")
    kpi_value_font = Font(name=_FONT, size=12, bold=True, color=_NAVY)
    kpi_label_font = Font(name=_FONT, size=9, color="718096")
    section_font = Font(name=_FONT, size=11, bold=True, color=_NAVY)
    section_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    alt_fill_1 = PatternFill(start_color=_ALT_ROW_1, end_color=_ALT_ROW_1, fill_type="solid")
    alt_fill_2 = PatternFill(start_color=_ALT_ROW_2, end_color=_ALT_ROW_2, fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color=_BORDER_COLOR),
        right=Side(style="thin", color=_BORDER_COLOR),
        top=Side(style="thin", color=_BORDER_COLOR),
        bottom=Side(style="thin", color=_BORDER_COLOR),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Grade fill cache
    grade_fills = {}
    grade_fonts = {}
    for grade, (bg, fg) in _GRADE_FILLS.items():
        grade_fills[grade] = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        grade_fonts[grade] = Font(name=_FONT, size=10, bold=True, color=fg)

    def _style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = center_align
            cell.border = thin_border

    def _write_row(ws, row, data, is_alt=False):
        fill = alt_fill_1 if is_alt else alt_fill_2
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = center_align
            if isinstance(val, float):
                cell.number_format = "0.0"

    def _auto_width(ws, min_w=10, max_w=40):
        """Auto-fit column widths based on cell content."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    # Account for merged cells
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
            width = min(max(max_len + 3, min_w), max_w)
            ws.column_dimensions[col_letter].width = width

    # ===== Summary Sheet =====
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = _NAVY

    # -- Title banner (merged navy bar) --
    banner_cols = 6
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=banner_cols)
    title_cell = ws["A1"]
    title_cell.value = f"GeoView QC Report"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    # Fill remaining cells in merged range
    for r in range(1, 3):
        for c in range(1, banner_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = navy_fill
            cell.border = Border()  # no border on banner

    # Subtitle row
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=banner_cols)
    sub_cell = ws["A3"]
    sub_cell.value = (
        f"{summary.project_name}  |  Client: {summary.client}  |  "
        f"Vessel: {summary.vessel}  |  {datetime.now():%Y-%m-%d %H:%M}"
    )
    info_bar_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    sub_cell.font = Font(name=_FONT, size=9, color="4A5568")
    sub_cell.fill = info_bar_fill
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, banner_cols + 1):
        ws.cell(row=3, column=c).fill = info_bar_fill

    # -- KPI cards row --
    kpi_row = 5
    ws.merge_cells(start_row=kpi_row, start_column=1, end_row=kpi_row, end_column=banner_cols)
    ws.cell(row=kpi_row, column=1, value="Key Performance Indicators").font = section_font

    kpis = [
        ("Total Files", str(summary.total_files)),
        ("Analyzed", str(summary.analyzed_files)),
        ("Avg Score", f"{summary.avg_score:.1f}"),
        ("PASS", str(summary.pass_count)),
        ("WARN", str(summary.warn_count)),
        ("FAIL", str(summary.fail_count)),
    ]

    label_row = kpi_row + 1
    value_row = kpi_row + 2
    for col, (label, value) in enumerate(kpis, 1):
        lc = ws.cell(row=label_row, column=col, value=label)
        lc.font = kpi_label_font
        lc.alignment = center_align

        vc = ws.cell(row=value_row, column=col, value=value)
        vc.font = kpi_value_font
        vc.alignment = center_align
        vc.border = Border(bottom=Side(style="medium", color=_NAVY))

    # -- Project details section --
    detail_start = value_row + 2
    ws.merge_cells(
        start_row=detail_start, start_column=1,
        end_row=detail_start, end_column=banner_cols,
    )
    ws.cell(row=detail_start, column=1, value="Project Details").font = section_font

    info_rows = [
        ("Project Name", summary.project_name),
        ("Client", summary.client),
        ("Vessel", summary.vessel),
        ("Report Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Overall Status", summary.overall_status.value),
        ("Score Range", f"{summary.min_score:.1f} – {summary.max_score:.1f}"),
        ("Critical Issues", str(summary.critical_issues)),
        ("Total Issues", str(summary.total_issues)),
    ]

    headers_row = detail_start + 1
    ws.cell(row=headers_row, column=1, value="Parameter").font = white_font
    ws.cell(row=headers_row, column=1).fill = navy_fill
    ws.cell(row=headers_row, column=1).border = thin_border
    ws.cell(row=headers_row, column=1).alignment = center_align
    ws.cell(row=headers_row, column=2, value="Value").font = white_font
    ws.cell(row=headers_row, column=2).fill = navy_fill
    ws.cell(row=headers_row, column=2).border = thin_border
    ws.cell(row=headers_row, column=2).alignment = center_align

    for i, (label, value) in enumerate(info_rows):
        r = headers_row + 1 + i
        is_alt = (i % 2 == 0)
        fill = alt_fill_1 if is_alt else alt_fill_2

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = header_font
        lc.border = thin_border
        lc.fill = fill
        lc.alignment = left_align

        vc = ws.cell(row=r, column=2, value=str(value))
        vc.font = body_font
        vc.border = thin_border
        vc.fill = fill
        vc.alignment = left_align

    _auto_width(ws)

    # ===== Per-domain result sheets =====
    domain_results: dict[str, list[QCResult]] = {}
    for r in summary.results:
        domain_results.setdefault(r.domain.value, []).append(r)

    for domain_name in ("mag", "sonar", "seismic"):
        results = domain_results.get(domain_name, [])
        ws = wb.create_sheet(title=f"{domain_name.upper()} QC")
        ws.sheet_properties.tabColor = _NAVY

        # Sheet title banner
        headers = ["File", "Line", "Score", "Grade", "Status", "Issues", "Duration (ms)"]
        col_count = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        tc = ws["A1"]
        tc.value = f"{domain_name.upper()} QC Results"
        tc.font = Font(name=_FONT, size=13, bold=True, color=_WHITE)
        tc.fill = navy_fill
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for c in range(1, col_count + 1):
            ws.cell(row=1, column=c).fill = navy_fill

        # Domain summary row
        if results:
            avg = sum(r.total_score for r in results) / len(results)
            pass_n = sum(1 for r in results if r.status == QCStatus.PASS)
            warn_n = sum(1 for r in results if r.status == QCStatus.WARN)
            fail_n = sum(1 for r in results if r.status == QCStatus.FAIL)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
            domain_info_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
            sc = ws["A2"]
            sc.value = (
                f"Files: {len(results)}  |  Avg Score: {avg:.1f}  |  "
                f"PASS: {pass_n}  WARN: {warn_n}  FAIL: {fail_n}"
            )
            sc.font = muted_font
            sc.fill = domain_info_fill
            for c in range(1, col_count + 1):
                ws.cell(row=2, column=c).fill = domain_info_fill

        # Header row
        header_row = 4
        for col, h in enumerate(headers, 1):
            ws.cell(row=header_row, column=col, value=h)
        _style_header(ws, header_row, col_count)

        # Data rows
        for i, r in enumerate(results):
            row_num = header_row + 1 + i
            row_data = [
                r.file_name,
                r.line_name,
                round(r.total_score, 1),
                r.grade.value,
                r.status.value,
                sum(r.issue_counts.values()),
                round(r.duration_ms, 0),
            ]
            _write_row(ws, row_num, row_data, is_alt=(i % 2 == 1))

            # Score color
            tier = _score_tier(r.total_score)
            score_cell = ws.cell(row=row_num, column=3)
            score_cell.font = Font(
                name=_FONT, size=10, bold=True, color=_SCORE_COLORS[tier],
            )

            # Grade cell with colored fill
            grade_cell = ws.cell(row=row_num, column=4)
            if r.grade in grade_fills:
                grade_cell.fill = grade_fills[r.grade]
                grade_cell.font = grade_fonts[r.grade]

            # Status cell coloring
            status_cell = ws.cell(row=row_num, column=5)
            if r.status == QCStatus.PASS:
                status_cell.font = Font(name=_FONT, size=10, bold=True, color="38A169")
            elif r.status == QCStatus.WARN:
                status_cell.font = Font(name=_FONT, size=10, bold=True, color="ED8936")
            elif r.status == QCStatus.FAIL:
                status_cell.font = Font(name=_FONT, size=10, bold=True, color="E53E3E")

            # File name left-aligned
            ws.cell(row=row_num, column=1).alignment = left_align
            ws.cell(row=row_num, column=2).alignment = left_align

        if not results:
            ws.cell(row=header_row + 1, column=1, value="No results").font = Font(
                name=_FONT, size=10, italic=True, color="718096",
            )

        _auto_width(ws)

    # Save
    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path
    else:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf


# ---------------------------------------------------------------------------
# Word Report
# ---------------------------------------------------------------------------

def generate_word_report(
    summary: QCProjectSummary,
    output_path: Optional[str | Path] = None,
) -> BytesIO | Path:
    """Generate unified Word QC report using GeoView design system.

    Features:
    - Professional title page with GeoView branding
    - Table of Contents placeholder
    - Proper Heading styles (H1/H2) for navigation
    - Styled tables with colored headers
    - Page numbers in footer
    - Pretendard font with Calibri fallback
    """
    from geoview_common.reporting.design_system import WordBuilder, WORD_STYLES

    style = WORD_STYLES["geoview_report"]
    wb = WordBuilder(style)

    # -- Cover page --
    wb.cover(
        title="GeoView QC Report",
        subtitle=summary.project_name,
        meta=(
            f"Client: {summary.client}  |  "
            f"Vessel: {summary.vessel}  |  "
            f"Date: {datetime.now():%Y-%m-%d}"
        ),
    )

    # -- Page footer with page numbers --
    wb.setup_page_footer(
        left_text=f"GeoView QC Report — {summary.project_name}",
    )

    wb.page_break()

    # -- Table of Contents placeholder --
    wb.heading("Table of Contents", level=1)
    wb.callout(
        "To generate the Table of Contents, right-click here in Word and "
        "select 'Update Field', or press Ctrl+A then F9.",
        kind="NOTE",
    )

    # Insert TOC field code
    try:
        from docx.oxml.ns import qn, nsdecls
        from docx.oxml import parse_xml
        p = wb.doc.add_paragraph()
        # TOC field: begin
        r1 = parse_xml(
            f'<w:r {nsdecls("w")}>'
            f'<w:fldChar w:fldCharType="begin"/></w:r>'
        )
        p._element.append(r1)
        # TOC instruction
        r2 = parse_xml(
            f'<w:r {nsdecls("w")}>'
            f'<w:instrText xml:space="preserve"> TOC \\o "1-3" \\h \\z \\u </w:instrText></w:r>'
        )
        p._element.append(r2)
        # TOC field: separate
        r3 = parse_xml(
            f'<w:r {nsdecls("w")}>'
            f'<w:fldChar w:fldCharType="separate"/></w:r>'
        )
        p._element.append(r3)
        # Placeholder text
        r4 = parse_xml(
            f'<w:r {nsdecls("w")}>'
            f'<w:rPr><w:i/><w:color w:val="718096"/></w:rPr>'
            f'<w:t>[Update this field to generate Table of Contents]</w:t></w:r>'
        )
        p._element.append(r4)
        # TOC field: end
        r5 = parse_xml(
            f'<w:r {nsdecls("w")}>'
            f'<w:fldChar w:fldCharType="end"/></w:r>'
        )
        p._element.append(r5)
    except Exception:
        pass  # graceful fallback if XML insertion fails

    wb.page_break()

    # -- 1. Executive Summary --
    wb.heading("Executive Summary", level=1)
    wb.body_text(
        f"This report presents the Quality Control analysis results for "
        f"the {summary.project_name} project. A total of {summary.total_files} "
        f"files were processed, with {summary.analyzed_files} files successfully "
        f"analyzed across all active domains. The overall average QC score is "
        f"{summary.avg_score:.1f} out of 100."
    )

    # Status summary
    status_text = (
        f"PASS: {summary.pass_count}  |  "
        f"WARN: {summary.warn_count}  |  "
        f"FAIL: {summary.fail_count}"
    )
    if summary.fail_count > 0:
        wb.callout(f"Overall Status: FAIL — {status_text}", kind="WARNING")
    elif summary.warn_count > 0:
        wb.callout(f"Overall Status: WARN — {status_text}", kind="NOTE")
    else:
        wb.callout(f"Overall Status: PASS — {status_text}", kind="NOTE")

    wb.page_break()

    # -- 2. Project Information --
    wb.heading("Project Information", level=1)
    wb.table(
        ["Parameter", "Value"],
        [
            ["Project Name", summary.project_name],
            ["Client", summary.client],
            ["Vessel", summary.vessel],
            ["Report Date", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Total Files", str(summary.total_files)],
            ["Analyzed Files", str(summary.analyzed_files)],
            ["Average Score", f"{summary.avg_score:.1f} / 100"],
            ["Score Range", f"{summary.min_score:.1f} – {summary.max_score:.1f}"],
            ["Overall Status", summary.overall_status.value],
            ["PASS / WARN / FAIL", f"{summary.pass_count} / {summary.warn_count} / {summary.fail_count}"],
            ["Critical Issues", str(summary.critical_issues)],
            ["Total Issues", str(summary.total_issues)],
        ],
    )

    # -- 3. Per-domain results --
    domain_results: dict[str, list[QCResult]] = {}
    for r in summary.results:
        domain_results.setdefault(r.domain.value, []).append(r)

    domain_labels = {
        "mag": ("MAG QC Results", "Magnetic"),
        "sonar": ("Sonar QC Results", "Sonar"),
        "seismic": ("Seismic QC Results", "Seismic"),
    }

    for domain_name in ("mag", "sonar", "seismic"):
        results = domain_results.get(domain_name, [])
        label, full_name = domain_labels[domain_name]

        wb.page_break()
        wb.heading(label, level=1)

        if not results:
            wb.callout(f"No {full_name} QC results available for this project.", kind="NOTE")
            continue

        # Domain overview
        avg = sum(r.total_score for r in results) / len(results)
        pass_n = sum(1 for r in results if r.status == QCStatus.PASS)
        warn_n = sum(1 for r in results if r.status == QCStatus.WARN)
        fail_n = sum(1 for r in results if r.status == QCStatus.FAIL)

        wb.body_text(
            f"{len(results)} files analyzed for {full_name} QC. "
            f"Average score: {avg:.1f}/100. "
            f"Results: {pass_n} PASS, {warn_n} WARN, {fail_n} FAIL."
        )

        # Results summary table
        wb.heading("Results Summary", level=2)
        headers = ["File", "Score", "Grade", "Status", "Issues"]
        rows = [
            [
                r.file_name,
                f"{r.total_score:.1f}",
                r.grade.value,
                r.status.value,
                str(sum(r.issue_counts.values())),
            ]
            for r in results
        ]
        wb.table(headers, rows)

        # Domain average callout
        kind = "WARNING" if fail_n > 0 else "NOTE"
        wb.callout(
            f"Average {full_name} Score: {avg:.1f} / 100  |  "
            f"PASS: {pass_n}  WARN: {warn_n}  FAIL: {fail_n}",
            kind=kind,
        )

        # Per-file detail (subsections for files with issues)
        files_with_issues = [r for r in results if sum(r.issue_counts.values()) > 0]
        if files_with_issues:
            wb.heading("Issue Details", level=2)
            for r in files_with_issues[:10]:  # Limit to first 10
                issue_total = sum(r.issue_counts.values())
                wb.body_text(
                    f"{r.file_name} — Score: {r.total_score:.1f}, "
                    f"Grade: {r.grade.value}, "
                    f"Issues: {issue_total} "
                    f"(Critical: {r.issue_counts.get('critical', 0)}, "
                    f"Warning: {r.issue_counts.get('warning', 0)})"
                )
            if len(files_with_issues) > 10:
                wb.callout(
                    f"Showing 10 of {len(files_with_issues)} files with issues. "
                    f"See Excel report for complete details.",
                    kind="NOTE",
                )

    # -- Footer --
    wb.hr()
    wb.footer(
        f"Generated by GeoView QC Suite v2.0  |  "
        f"{datetime.now():%Y-%m-%d %H:%M}  |  "
        f"\u00A9 2025-2026 Geoview Co., Ltd."
    )

    # Save
    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path
    else:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)
        buf = BytesIO(Path(tmp_path).read_bytes())
        Path(tmp_path).unlink(missing_ok=True)
        return buf


# ---------------------------------------------------------------------------
# PDF Report
# ---------------------------------------------------------------------------

def generate_pdf_report(
    summary: QCProjectSummary,
    output_path: Optional[str | Path] = None,
) -> BytesIO | Path:
    """Generate unified PDF QC report using reportlab.

    Features:
    - GeoView navy header bar on every page
    - Page numbers in footer
    - Alternating row colors in tables
    - Score coloring (green/orange/red)
    - Professional cover page
    - Copyright footer
    """
    try:
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, KeepTogether,
        )
        from reportlab.platypus.frames import Frame
    except ImportError:
        raise ImportError("reportlab required: pip install reportlab")

    NAVY = rl_colors.HexColor("#1E3A5F")
    NAVY_LIGHT = rl_colors.HexColor("#2D5F8A")
    LIGHT = rl_colors.HexColor("#F7FAFC")
    BORDER = rl_colors.HexColor("#CBD5E0")
    GREEN = rl_colors.HexColor("#38A169")
    ORANGE = rl_colors.HexColor("#ED8936")
    RED = rl_colors.HexColor("#E53E3E")
    MUTED = rl_colors.HexColor("#718096")
    SECTION_BG = rl_colors.HexColor("#EDF2F7")

    # Grade colors for PDF
    grade_colors = {
        "A": rl_colors.HexColor("#C6F6D5"),
        "B": rl_colors.HexColor("#BEE3F8"),
        "C": rl_colors.HexColor("#FEEBC8"),
        "D": rl_colors.HexColor("#FED7D7"),
        "F": rl_colors.HexColor("#E2E8F0"),
    }

    buf = BytesIO()
    target = str(output_path) if output_path else buf

    # Page header/footer callback
    _report_title = f"GeoView QC Report — {summary.project_name}"
    _copyright = f"\u00A9 2025-2026 Geoview Co., Ltd."
    _gen_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    def _header_footer(canvas, doc):
        """Draw header bar and footer on every page."""
        canvas.saveState()
        w, h = A4

        # Header bar — full-width navy rectangle
        bar_h = 14 * mm
        canvas.setFillColor(NAVY)
        canvas.rect(0, h - bar_h, w, bar_h, fill=True, stroke=False)

        # Header text
        canvas.setFillColor(rl_colors.white)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawString(20 * mm, h - bar_h + 4.5 * mm, "GeoView QC Report")

        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(
            w - 20 * mm, h - bar_h + 4.5 * mm,
            summary.project_name,
        )

        # Footer line
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.5)
        canvas.line(20 * mm, 15 * mm, w - 20 * mm, 15 * mm)

        # Footer left: copyright
        canvas.setFillColor(MUTED)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(20 * mm, 10 * mm, f"{_copyright}  |  {_gen_time}")

        # Footer right: page number
        canvas.drawRightString(
            w - 20 * mm, 10 * mm,
            f"Page {doc.page}",
        )

        canvas.restoreState()

    doc = SimpleDocTemplate(
        target, pagesize=A4,
        topMargin=28 * mm,   # extra room for header bar
        bottomMargin=22 * mm,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        title=_report_title,
        author="GeoView QC Suite",
    )

    styles = getSampleStyleSheet()

    # Custom styles
    styles.add(ParagraphStyle(
        "CoverTitle", parent=styles["Title"],
        fontSize=28, textColor=NAVY, spaceAfter=4 * mm,
        fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        "CoverSubtitle", parent=styles["Heading2"],
        fontSize=16, textColor=NAVY_LIGHT, spaceAfter=6 * mm,
    ))
    styles.add(ParagraphStyle(
        "CoverMeta", parent=styles["Normal"],
        fontSize=10, textColor=MUTED, spaceAfter=2 * mm,
    ))
    styles.add(ParagraphStyle(
        "SectionHeading", parent=styles["Heading1"],
        fontSize=14, textColor=NAVY, fontName="Helvetica-Bold",
        spaceBefore=8 * mm, spaceAfter=4 * mm,
    ))
    styles.add(ParagraphStyle(
        "SubHeading", parent=styles["Heading2"],
        fontSize=11, textColor=NAVY_LIGHT, fontName="Helvetica-Bold",
        spaceBefore=6 * mm, spaceAfter=3 * mm,
    ))
    styles.add(ParagraphStyle(
        "GVBodyText", parent=styles["Normal"],
        fontSize=9, textColor=rl_colors.HexColor("#1A202C"),
        spaceAfter=3 * mm, leading=14,
    ))
    styles.add(ParagraphStyle(
        "FooterNote", parent=styles["Normal"],
        fontSize=7.5, textColor=MUTED, alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        "ScoreGreen", parent=styles["Normal"],
        fontSize=9, textColor=GREEN, fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        "ScoreOrange", parent=styles["Normal"],
        fontSize=9, textColor=ORANGE, fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        "ScoreRed", parent=styles["Normal"],
        fontSize=9, textColor=RED, fontName="Helvetica-Bold",
    ))

    def _score_para(score_val: float) -> Paragraph:
        """Return a Paragraph with colored score text."""
        tier = _score_tier(score_val)
        style_name = {
            "pass": "ScoreGreen",
            "warn": "ScoreOrange",
            "fail": "ScoreRed",
        }[tier]
        return Paragraph(f"{score_val:.1f}", styles[style_name])

    def _styled_table(data, col_widths, has_grade_col=False, grade_col_idx=None):
        """Create a professional styled table."""
        t = Table(data, colWidths=col_widths)
        style_cmds = [
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Body
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
            # Alternating rows
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, LIGHT]),
            # Padding
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]

        # Grade column coloring
        if has_grade_col and grade_col_idx is not None:
            for row_idx in range(1, len(data)):
                grade_val = data[row_idx][grade_col_idx]
                grade_str = grade_val if isinstance(grade_val, str) else str(grade_val)
                if grade_str in grade_colors:
                    style_cmds.append(
                        ("BACKGROUND", (grade_col_idx, row_idx),
                         (grade_col_idx, row_idx), grade_colors[grade_str])
                    )

        t.setStyle(TableStyle(style_cmds))
        return t

    elements: list = []

    # ===== Cover Page =====
    elements.append(Spacer(1, 35 * mm))
    elements.append(Paragraph("GeoView QC Report", styles["CoverTitle"]))
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph(summary.project_name, styles["CoverSubtitle"]))
    elements.append(Spacer(1, 8 * mm))

    # Cover info table
    cover_info = [
        ["Client", summary.client],
        ["Vessel", summary.vessel],
        ["Report Date", _gen_time],
        ["Total Files", str(summary.total_files)],
        ["Overall Status", summary.overall_status.value],
    ]
    cover_t = Table(cover_info, colWidths=[120, 280])
    cover_t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), NAVY),
        ("TEXTCOLOR", (1, 0), (1, -1), rl_colors.HexColor("#1A202C")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, BORDER),
        ("LINEBELOW", (0, -1), (-1, -1), 1, NAVY),
    ]))
    elements.append(cover_t)
    elements.append(PageBreak())

    # ===== Project Summary =====
    elements.append(Paragraph("Project Summary", styles["SectionHeading"]))

    summary_data = [
        ["Metric", "Value"],
        ["Project Name", summary.project_name],
        ["Client", summary.client],
        ["Vessel", summary.vessel],
        ["Total Files", str(summary.total_files)],
        ["Analyzed Files", str(summary.analyzed_files)],
        ["Average Score", f"{summary.avg_score:.1f}"],
        ["Score Range", f"{summary.min_score:.1f} \u2013 {summary.max_score:.1f}"],
        ["PASS / WARN / FAIL", f"{summary.pass_count} / {summary.warn_count} / {summary.fail_count}"],
        ["Critical Issues", str(summary.critical_issues)],
        ["Overall Status", summary.overall_status.value],
    ]

    elements.append(_styled_table(summary_data, [180, 220]))
    elements.append(Spacer(1, 5 * mm))

    # Status KPI bar
    kpi_data = [
        ["PASS", "WARN", "FAIL", "Avg Score"],
        [str(summary.pass_count), str(summary.warn_count),
         str(summary.fail_count), f"{summary.avg_score:.1f}"],
    ]
    kpi_t = Table(kpi_data, colWidths=[100, 100, 100, 100])
    kpi_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("TOPPADDING", (0, 1), (-1, 1), 8),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
        # Color the KPI values
        ("TEXTCOLOR", (0, 1), (0, 1), GREEN),
        ("TEXTCOLOR", (1, 1), (1, 1), ORANGE),
        ("TEXTCOLOR", (2, 1), (2, 1), RED),
        ("TEXTCOLOR", (3, 1), (3, 1), NAVY),
        ("BACKGROUND", (0, 1), (0, 1), rl_colors.HexColor("#F0FFF4")),
        ("BACKGROUND", (1, 1), (1, 1), rl_colors.HexColor("#FFFBEB")),
        ("BACKGROUND", (2, 1), (2, 1), rl_colors.HexColor("#FFF5F5")),
        ("BACKGROUND", (3, 1), (3, 1), LIGHT),
    ]
    kpi_t.setStyle(TableStyle(kpi_style_cmds))
    elements.append(kpi_t)

    # ===== Per-domain tables =====
    domain_results: dict[str, list[QCResult]] = {}
    for r in summary.results:
        domain_results.setdefault(r.domain.value, []).append(r)

    domain_labels = {
        "mag": "MAG QC Results",
        "sonar": "Sonar QC Results",
        "seismic": "Seismic QC Results",
    }

    for domain in ("mag", "sonar", "seismic"):
        results = domain_results.get(domain, [])
        elements.append(Spacer(1, 8 * mm))
        elements.append(Paragraph(domain_labels[domain], styles["SectionHeading"]))

        if not results:
            elements.append(Paragraph(
                f"No {domain.upper()} results available.",
                styles["GVBodyText"],
            ))
            continue

        # Domain summary text
        avg = sum(r.total_score for r in results) / len(results)
        pass_n = sum(1 for r in results if r.status == QCStatus.PASS)
        warn_n = sum(1 for r in results if r.status == QCStatus.WARN)
        fail_n = sum(1 for r in results if r.status == QCStatus.FAIL)
        elements.append(Paragraph(
            f"{len(results)} files analyzed  |  Average: {avg:.1f}/100  |  "
            f"PASS: {pass_n}  WARN: {warn_n}  FAIL: {fail_n}",
            styles["GVBodyText"],
        ))

        # Results table with colored scores
        data = [["File", "Score", "Grade", "Status", "Issues"]]
        for r in results:
            data.append([
                r.file_name,
                _score_para(r.total_score),
                r.grade.value,
                r.status.value,
                str(sum(r.issue_counts.values())),
            ])

        t = _styled_table(
            data,
            col_widths=[160, 60, 50, 55, 50],
            has_grade_col=True,
            grade_col_idx=2,
        )
        # Left-align file name column
        t.setStyle(TableStyle([
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ]))
        elements.append(t)

    # ===== Final footer =====
    elements.append(Spacer(1, 15 * mm))
    elements.append(Paragraph(
        f"Generated by GeoView QC Suite v2.0  |  {_gen_time}  |  {_copyright}",
        styles["FooterNote"],
    ))

    doc.build(elements, onFirstPage=_header_footer, onLaterPages=_header_footer)

    if output_path:
        return Path(output_path)
    buf.seek(0)
    return buf
